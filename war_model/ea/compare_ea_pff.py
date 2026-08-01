"""Side-by-side workbook: the EA-rated WAR against the PFF-rated one, 2026.

Both numbers are wins above replacement on the same wins map and the same inherited
concept weights, so they are directly comparable. What differs is the measurement
layer underneath - PFF grades and CFBD play value on one side, EA attribute ratings on
the other - and therefore what each one is willing to say about a player nobody has
watched yet.

NEITHER COLUMN IS SCORED HERE, because there is nothing to score against. 2026 has not
been played, and CFB 27 cannot be run retrospectively on 2025 (it rates 2026 rosters,
so 2025's seniors are missing and the 400+ snap players match at 53%). This shows where
the two disagree, not which is right.

Run: ../../venv/bin/python compare_ea_pff.py
"""
from __future__ import annotations

import os
import sys

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
WAR_DIR = os.path.dirname(HERE)
sys.path.insert(0, WAR_DIR)
from build_roster_2026 import norm_name  # noqa: E402

OUT = f"{HERE}/ea_vs_pff_2026.xlsx"
HDR = PatternFill("solid", fgColor="1F3B4D")
POS = PatternFill("solid", fgColor="D6EAD6")
NEG = PatternFill("solid", fgColor="F5D6D0")
THIN = Border(bottom=Side(style="thin", color="BBBBBB"))


def load():
    ea_t = pd.read_csv(f"{HERE}/ea_team_war_2026.csv")
    ea_p = pd.read_csv(f"{HERE}/ea_player_war_2026.csv")
    pff = pd.read_csv(f"{WAR_DIR}/projections_2026_v2.csv")
    pff["key"] = pff.player.map(norm_name)
    return ea_t, ea_p, pff


def team_sheet(ea_t, pff):
    ours = pff.groupby("team", as_index=False).proj_war.sum().rename(
        columns={"proj_war": "pff_war"})
    d = ours.merge(ea_t[["team", "war"]].rename(columns={"war": "ea_war"}),
                   on="team", how="inner")
    # both are wins above replacement but the two builds do not have to sum to the
    # same league total, so ranks are compared on each scale's own terms
    d["pff_rank"] = d.pff_war.rank(ascending=False, method="min").astype(int)
    d["ea_rank"] = d.ea_war.rank(ascending=False, method="min").astype(int)
    d["rank_move"] = d.pff_rank - d.ea_rank          # + = EA likes them more
    d["war_diff"] = d.ea_war - d.pff_war
    d["z_diff"] = ((d.ea_war - d.ea_war.mean()) / d.ea_war.std()
                   - (d.pff_war - d.pff_war.mean()) / d.pff_war.std())
    return d.sort_values("pff_rank")[
        ["team", "pff_rank", "ea_rank", "rank_move", "pff_war", "ea_war",
         "war_diff", "z_diff"]]


def player_sheet(ea_p, pff):
    # EA's own group column is dropped: pff.broad_group is renamed to `group` below,
    # and two columns of that name make the frame un-groupable.
    e = ea_p[["key", "team", "player", "position", "overall", "war"]].rename(
        columns={"war": "ea_war", "overall": "ea_ovr", "position": "ea_pos"})
    e = e.drop_duplicates(["key", "team"])
    d = pff.merge(e, on=["key", "team"], how="left", suffixes=("", "_ea"))
    d["war_diff"] = d.ea_war - d.proj_war
    return d.rename(columns={"proj_war": "pff_war", "broad_group": "group",
                             "roster_position": "pos"})[
        ["player", "team", "conference", "pos", "group", "class", "depth",
         "is_starter", "imputed", "snaps_2025", "war_2025", "pff_war",
         "ea_ovr", "ea_war", "war_diff"]]


def group_sheet(pl):
    g = pl.dropna(subset=["ea_war"]).groupby("group").agg(
        players=("player", "size"),
        pff_mean=("pff_war", "mean"), ea_mean=("ea_war", "mean"),
        pff_total=("pff_war", "sum"), ea_total=("ea_war", "sum"),
        corr=("pff_war", lambda s: np.nan)).reset_index()
    cors = pl.dropna(subset=["ea_war"]).groupby("group").apply(
        lambda x: x.pff_war.corr(x.ea_war), include_groups=False)
    g["corr"] = g.group.map(cors)
    g["mean_diff"] = g.ea_mean - g.pff_mean
    g["share_pff"] = g.pff_total / g.pff_total.sum()
    g["share_ea"] = g.ea_total / g.ea_total.sum()
    return g.sort_values("share_pff", ascending=False)[
        ["group", "players", "corr", "pff_mean", "ea_mean", "mean_diff",
         "share_pff", "share_ea"]]


def validation_sheet():
    """The one real out-of-sample test available, and it is team-level.

    collegefootball.gg published CFB 26's per-team ratings on 2025-07-06, before that
    game's launch week and before a snap of the 2025 season. That is a clean entering-
    2025 EA signal, and 2025 has been played - so for once both sides can be scored on
    the same outcome with no leakage on either.

    It tests EA's ratings, not the full EA WAR pipeline: these are EA's own team
    aggregates rather than the player-level build. A properly constructed EA WAR could
    in principle extract more than EA's own summary does.
    """
    import numpy as np
    ea = pd.read_csv(f"{HERE}/cfb26_team.csv")
    ea["team"] = ea.team.replace({"Appalachian State": "App State",
                                  "San Jose State": "San José State",
                                  "FIU": "Florida International", "Hawaii": "Hawai'i"})
    ea = ea.set_index("team")
    recs = pd.read_csv(f"{WAR_DIR}/records.csv")
    y = recs[recs.season == 2025].set_index("team").adj_win_pct

    sys.path.insert(0, os.path.dirname(WAR_DIR))
    from scripts.train import load_bundle, blended_talent
    from src.data import pff
    std, cfbd_tal, ret, games, pyth = load_bundle()
    ours = blended_talent(cfbd_tal, pff.build_roster_talent())[2025]

    idx = ea.index.intersection(y.index).intersection(ours.index)
    z = lambda s: (s - s.mean()) / s.std()
    w = y.reindex(idx)
    rows = [
        ("n teams", len(idx)),
        ("EA CFB26 overall -> 2025 adj win pct", round(float(ea.ovr.reindex(idx).corr(w)), 3)),
        ("EA CFB26 offense+defense -> same",
         round(float(((z(ea.off.reindex(idx)) + z(ea["def"].reindex(idx))) / 2).corr(w)), 3)),
        ("EA CFB26 prestige -> same", round(float(ea.prestige.reindex(idx).corr(w)), 3)),
        ("OUR talent (PFF+CFBD+WAR) -> same", round(float(ours.reindex(idx).corr(w)), 3)),
    ]
    d = pd.DataFrame({"ea": z(ea.ovr.reindex(idx)), "ours": z(ours.reindex(idx)),
                      "y": w}).dropna()
    X = np.column_stack([np.ones(len(d)), d.ours])
    b = np.linalg.lstsq(X, d.y, rcond=None)[0]
    res = d.y - X @ b
    X2 = np.column_stack([np.ones(len(d)), d.ours, d.ea])
    b2 = np.linalg.lstsq(X2, d.y, rcond=None)[0]
    rows += [
        ("", ""),
        ("EA vs the residual after our talent", round(float(d.ea.corr(res)), 3)),
        ("R2, our talent alone", round(float(1 - res.var() / d.y.var()), 3)),
        ("R2, our talent + EA", round(float(1 - (d.y - X2 @ b2).var() / d.y.var()), 3)),
        ("", ""),
        ("VERDICT", "EA is the weaker signal on its own (.42 vs .51) and adds "
                    "essentially nothing on top of what we already have (+.004 R2, "
                    "residual correlation -.03)."),
        ("CAVEAT", "Team-level only. These are EA's own aggregates, not a player-level "
                   "EA WAR, which could extract more from the same ratings."),
    ]
    return pd.DataFrame(rows, columns=["test", "value"])


def readme(team, pl):
    cov = pl.ea_war.notna().mean()
    imp = pl[pl.imputed]
    r = team.pff_war.corr(team.ea_war)
    rows = [
        ("WHAT THIS IS", ""),
        ("", "Two builds of player wins-above-replacement for 2026, on the same wins "
             "map and the same concept weights. Only the measurement layer differs."),
        ("pff_war", "The current model. PFF grades + CFBD play value -> facets -> "
                    "concepts -> WAR, projected to 2026 by a gradient-boosted model."),
        ("ea_war", "The same machinery with EA College Football 27 launch ratings as "
                   "the measurement layer instead. Concept weights are INHERITED from "
                   "the PFF fit (11 seasons, 1,302 team-seasons)."),
        ("", ""),
        ("WHAT IT DOES NOT SAY", ""),
        ("", "Neither column is scored. 2026 has not been played, so there is no "
             "outcome to check either against."),
        ("", "EA could not be validated on a past season either. CFB 27 rates 2026 "
             "rosters, so 2025's seniors are absent - the 400+ snap players match at "
             "53%, worse than the scrubs."),
        ("", "CFB 25 and 26 are unobtainable: EA serves only the current game, CFB 26 "
             "sits behind a robots.txt disallowing this crawler, and no archive holds "
             "the player payload. So EA weights could not be fitted, only borrowed."),
        ("", ""),
        ("WHERE THEY AGREE", ""),
        ("team-level correlation", round(float(r), 3)),
        ("player-level coverage", f"{cov*100:.1f}% of our 5,770 slots have an EA rating"),
        ("", ""),
        ("THE ONE CLEAR EA ADVANTAGE", ""),
        ("our imputed slots", f"{len(imp)} players currently get a positional prior "
                              f"rather than a measurement"),
        ("of those, EA rates", f"{imp.ea_war.notna().mean()*100:.1f}%"),
        ("", "That is the gap EA fills that PFF structurally cannot: it has an opinion "
             "on players who have never taken a snap."),
        ("", ""),
        ("WHERE THEY DISAGREE, AND WHY IT MATTERS", ""),
        ("", "The disagreement is not noise. EA down-rates every service academy "
             "massively - Army 24th on our model against 111th on EA's, Navy 37 -> 86, "
             "Air Force 49 -> 84 - because the flexbone runs on undersized, lightly "
             "recruited players whose individual attribute ratings are poor and whose "
             "collective production is not."),
        ("", "That is the failure mode of an editorial rating in one picture: EA "
             "measures how good the players look, our model measures what the unit "
             "actually did. Where those come apart - option teams, heavy scheme "
             "identity, developmental programs - EA will be wrong in a predictable "
             "direction."),
        ("recruiting tilt", "corr(disagreement, recruiting composite) = +0.24. Top "
                            "recruiting quartile averages +2.9 ranks better under EA, "
                            "bottom quartile -3.1."),
        ("", ""),
        ("KNOWN WEAKNESSES OF THE EA COLUMN", ""),
        ("editorial, not measured", "EA team talent correlates 0.80-0.85 with the "
                                    "recruiting composite, against 0.38 for PFF talent "
                                    "and 0.53 for WAR. 88% of it is spanned by "
                                    "recruiting + our existing WAR."),
        ("no penalties", "EA rates no discipline equivalent - 6.0% of fitted weight, "
                         "redistributed across the other concepts."),
        ("no playing time", "EA rates everyone equally. Depth here is EA's own overall "
                            "rank within team and position; snaps come from what a "
                            "player of that rank actually played in 2025."),
    ]
    return pd.DataFrame(rows, columns=["field", "value"])


def style(ws, df, money_cols=(), diff_cols=(), width_overrides=None):
    ws.freeze_panes = "A2"
    for c in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill, cell.font = HDR, Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    widths = width_overrides or {}
    for i, col in enumerate(df.columns, start=1):
        L = get_column_letter(i)
        ws.column_dimensions[L].width = widths.get(col, max(10, min(22, len(col) + 4)))
    dcols = [df.columns.get_loc(c) + 1 for c in diff_cols if c in df.columns]
    for r in range(2, len(df) + 2):
        for c in dcols:
            v = ws.cell(row=r, column=c).value
            if isinstance(v, (int, float)):
                ws.cell(row=r, column=c).fill = POS if v > 0 else (NEG if v < 0 else PatternFill())
        for c in [df.columns.get_loc(x) + 1 for x in money_cols if x in df.columns]:
            ws.cell(row=r, column=c).number_format = "0.000"


def main():
    ea_t, ea_p, pff = load()
    team = team_sheet(ea_t, pff)
    pl = player_sheet(ea_p, pff)
    grp = group_sheet(pl)
    info = readme(team, pl)

    disagree = pl.dropna(subset=["ea_war"]).reindex(
        pl.dropna(subset=["ea_war"]).war_diff.abs().sort_values(ascending=False).index)
    imputed = pl[pl.imputed].sort_values("ea_war", ascending=False)

    val = validation_sheet()

    with pd.ExcelWriter(OUT, engine="openpyxl") as xl:
        info.to_excel(xl, sheet_name="Read me", index=False)
        val.to_excel(xl, sheet_name="2025 validation", index=False)
        team.to_excel(xl, sheet_name="Teams", index=False)
        team.reindex(team.rank_move.abs().sort_values(ascending=False).index
                     ).to_excel(xl, sheet_name="Teams by disagreement", index=False)
        pl.sort_values("pff_war", ascending=False).to_excel(
            xl, sheet_name="Players", index=False)
        disagree.head(500).to_excel(xl, sheet_name="Biggest disagreements", index=False)
        imputed.to_excel(xl, sheet_name="Our imputed players", index=False)
        grp.to_excel(xl, sheet_name="By position group", index=False)

        style(xl.sheets["Read me"], info, width_overrides={"field": 26, "value": 110})
        xl.sheets["Read me"].column_dimensions["B"].width = 110
        for r in range(2, len(info) + 2):
            xl.sheets["Read me"].cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
            xl.sheets["Read me"].cell(row=r, column=1).font = Font(bold=True)
        for nm, d in (("Teams", team), ("Teams by disagreement", team)):
            style(xl.sheets[nm], d, money_cols=["pff_war", "ea_war", "war_diff", "z_diff"],
                  diff_cols=["rank_move", "war_diff", "z_diff"], width_overrides={"team": 24})
        for nm, d in (("Players", pl), ("Biggest disagreements", disagree),
                      ("Our imputed players", imputed)):
            style(xl.sheets[nm], d, money_cols=["pff_war", "ea_war", "war_diff", "war_2025"],
                  diff_cols=["war_diff"],
                  width_overrides={"player": 24, "team": 20, "conference": 20})
        style(xl.sheets["2025 validation"], val,
              width_overrides={"test": 42, "value": 100})
        for r in range(2, len(val) + 2):
            xl.sheets["2025 validation"].cell(row=r, column=1).font = Font(bold=True)
            xl.sheets["2025 validation"].cell(row=r, column=2).alignment = Alignment(
                wrap_text=True, vertical="top")
        style(xl.sheets["By position group"], grp,
              money_cols=["corr", "pff_mean", "ea_mean", "mean_diff", "share_pff", "share_ea"],
              diff_cols=["mean_diff"])

    print(f"-> {OUT}")
    print(f"\nteam-level correlation pff_war ~ ea_war: {team.pff_war.corr(team.ea_war):.3f}")
    print(f"player-level, where both exist:           "
          f"{pl.pff_war.corr(pl.ea_war):.3f}  (n={int(pl.ea_war.notna().sum())})")
    print("\nteams EA likes most (rank_move = pff_rank - ea_rank):")
    print(team.nlargest(8, "rank_move")[["team", "pff_rank", "ea_rank", "rank_move"]]
          .to_string(index=False))
    print("\nteams EA likes least:")
    print(team.nsmallest(8, "rank_move")[["team", "pff_rank", "ea_rank", "rank_move"]]
          .to_string(index=False))


if __name__ == "__main__":
    main()
